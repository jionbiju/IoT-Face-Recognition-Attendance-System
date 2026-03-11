"""
Generate a sample Excel file in college format to demonstrate the output
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def generate_sample_excel():
    """Generate a sample Excel file showing the college format"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CST362"
    
    # Sample data
    students = [
        (1, "Student A", "1"),
        (2, "Student B", "2"),
        (3, "Student C", "3"),
        (4, "Student D", "4"),
        (5, "Student E", "5"),
    ]
    
    # Sample dates (last 10 days)
    dates = []
    for i in range(10, 0, -1):
        date = datetime.date.today() - datetime.timedelta(days=i)
        dates.append(date.strftime("%Y-%m-%d"))
    
    # Sample attendance (random pattern)
    # student_id -> {date: is_present}
    attendance_map = {
        1: {dates[0]: True, dates[2]: True, dates[4]: True, dates[6]: True, dates[8]: True},
        2: {dates[1]: True, dates[3]: True, dates[5]: True, dates[7]: True, dates[9]: True},
        3: {dates[0]: True, dates[1]: True, dates[2]: True, dates[3]: True, dates[4]: True},
        4: {dates[5]: True, dates[6]: True, dates[7]: True, dates[8]: True, dates[9]: True},
        5: {dates[0]: True, dates[2]: True, dates[4]: True, dates[6]: True, dates[8]: True},
    }
    
    # --- HEADER SECTION ---
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "Attendance Summary"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:C2')
    subject_cell = ws['A2']
    subject_cell.value = "Subject: CST362 - PROGRAMMING IN PYTHON"
    subject_cell.font = Font(bold=True, size=11)
    
    ws.merge_cells('A3:C3')
    date_cell = ws['A3']
    date_cell.value = f"Period: {dates[0]} to {dates[-1]}"
    date_cell.font = Font(size=10)
    
    # Empty row
    ws.append([])
    
    # --- DATA TABLE HEADER ---
    header_row = 5
    
    ws.cell(row=header_row, column=1, value="Roll No")
    ws.cell(row=header_row, column=2, value="Student")
    
    # Date columns
    date_col_start = 3
    for idx, date_str in enumerate(dates, start=date_col_start):
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        formatted_date = date_obj.strftime('%d/%m')
        period_num = idx - date_col_start + 1
        
        cell = ws.cell(row=header_row, column=idx, value=formatted_date)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Period number
        ws.cell(row=header_row + 1, column=idx, value=period_num)
        ws.cell(row=header_row + 1, column=idx).font = Font(size=8, italic=True)
        ws.cell(row=header_row + 1, column=idx).alignment = Alignment(horizontal='center')
    
    # Style header
    for col in range(1, len(dates) + 3):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # --- DATA ROWS ---
    data_row_start = header_row + 2
    
    for row_idx, (student_id, student_name, roll_no) in enumerate(students, start=data_row_start):
        ws.cell(row=row_idx, column=1, value=roll_no)
        ws.cell(row=row_idx, column=2, value=student_name)
        
        # Attendance for each date
        for col_idx, date_str in enumerate(dates, start=date_col_start):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            is_present = (student_id in attendance_map and 
                         date_str in attendance_map[student_id])
            
            if is_present:
                # Present = blank
                cell.value = ""
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                # Absent = 'A'
                cell.value = "A"
                cell.font = Font(bold=True, color="FF0000")
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Style student columns
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # --- COLUMN WIDTHS ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    
    for col_idx in range(date_col_start, len(dates) + date_col_start):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6
    
    # Save file
    filename = "SAMPLE_College_Format_Attendance.xlsx"
    wb.save(filename)
    
    print("=" * 80)
    print("✅ SAMPLE EXCEL FILE GENERATED")
    print("=" * 80)
    print()
    print(f"File created: {filename}")
    print()
    print("This file demonstrates the COLLEGE FORMAT with:")
    print("  ✓ Grid layout (dates as columns)")
    print("  ✓ Students as rows")
    print("  ✓ 'A' for Absent (red background)")
    print("  ✓ Blank for Present")
    print("  ✓ Professional formatting")
    print()
    print("Open this file in Excel to see the correct format!")
    print("=" * 80)

if __name__ == "__main__":
    generate_sample_excel()
